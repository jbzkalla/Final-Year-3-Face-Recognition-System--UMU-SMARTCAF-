import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from datetime import datetime

from utils.constants import BASE_DIR

def generate_excel_with_logo(data_rows, headers, title="Report", filename="report.xlsx"):
    """
    Generates an Excel file with the UMU logo at the top using openpyxl.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # 1. Add Logo
    logo_path = os.path.join(BASE_DIR, "umu_logo.jpeg")
    if os.path.exists(logo_path):
        try:
            img = Image(logo_path)
            # Resize image to fit roughly 3-4 rows
            img.width = 100
            img.height = 100
            
            # Position logo in A1:A4 area and center it
            # Using anchor for fixed positioning
            from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
            from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
            
            # Calculate offsets for centering (assuming col A width is roughly 15 units ~ 120 pixels)
            # Row height is usually 15 points ~ 20 pixels. 4 rows ~ 80-100 pixels.
            ws.column_dimensions['A'].width = 20
            for i in range(1, 6):
                ws.row_dimensions[i].height = 25
                
            img.anchor = 'A1' # Simple anchor for now, but we can refine with offsets
            ws.add_image(img)
            
            # Ensure image stays with cells
            # img.move_with_cells = True # Standard behavior often enough, but let's be explicit
            
            # Merge cells for title area - adjusted to start from B
            ws.merge_cells('B1:G2')
            cell = ws['B1']
            cell.value = "UGANDA MARTYRS UNIVERSITY"
            cell.font = Font(size=18, bold=True, color="610A0A")
            cell.alignment = Alignment(horizontal='center', vertical='center')

            ws.merge_cells('B3:G4')
            cell = ws['B3']
            cell.value = "SmartCaf System - Professional Report"
            cell.font = Font(size=14, italic=True, color="0858c7")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        except Exception as e:
            print(f"Logo error: {e}")

    # 2. Add Report Metadata
    start_row = 6
    ws.cell(row=start_row, column=1, value=title).font = Font(size=14, bold=True)
    ws.cell(row=start_row+1, column=1, value=f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # 3. Add Table Headers
    table_start_row = start_row + 3
    header_fill = PatternFill(start_color="610A0A", end_color="610A0A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=table_start_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        
        # Adjust column width
        ws.column_dimensions[ws.cell(row=table_start_row, column=col).column_letter].width = max(len(str(header)), 15)

    # 4. Add Data Rows
    for r_idx, row_data in enumerate(data_rows, 1):
        for c_idx, value in enumerate(row_data, 1):
            ws.cell(row=table_start_row + r_idx, column=c_idx, value=value)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
